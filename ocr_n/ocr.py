import os
import sys
import json
from pathlib import Path

import fitz  # PyMuPDF
from google import genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

API_KEY = os.getenv("GEMINI_API_KEY")

if not API_KEY:
    print("ERROR: GEMINI_API_KEY is not set.")
    print()
    print('Run:')
    print('$env:GEMINI_API_KEY="YOUR_API_KEY"')
    sys.exit(1)

client = genai.Client(api_key=API_KEY)


# ============================================================
# CLEAN NORMAL TEXT
# ============================================================

def clean_text(value):
    """Clean normal extracted text."""

    if value is None:
        return ""

    if isinstance(value, list):
        value = " ".join(str(x) for x in value)

    return " ".join(str(value).split()).strip()


# ============================================================
# CLEAN COMMENT
# ============================================================

def clean_comment(value):
    """
    Clean the Old:New:Comment field.

    Empty representations such as [][][] are removed.
    Real comments are preserved.
    """

    if value is None:
        return ""

    value = str(value).strip()

    empty_values = {
        "",
        "[]",
        "[][]",
        "[][][]",
        "[]:[]:[]",
        "[ ]:[ ]:[ ]",
        "[][ ][]",
    }

    if value in empty_values:
        return ""

    # Check whether the value contains only
    # brackets, colons and spaces.
    test = (
        value
        .replace("[", "")
        .replace("]", "")
        .replace(":", "")
        .replace(" ", "")
    )

    if test == "":
        return ""

    return value


# ============================================================
# EXTRACT ONE PAGE
# ============================================================

def extract_page(image_bytes, page_number):

    print(f"Processing page {page_number}...")

    prompt = r"""
You are an OCR and table extraction system.

The image is one page from a scanned audit-trail PDF.

Extract ALL audit trail records visible on this page.

The table columns are:

1. Event Date/Time
2. Accession Number
3. Event Type/Description
4. Event Associated User ID
5. Old:New:Comment
6. Card Bar Code

IMPORTANT:

- Extract every visible record.
- Do not skip rows.
- Do not summarize.
- Do not invent information.
- Preserve the original information as accurately as possible.
- If a field is empty, return an empty string.

SPECIAL RULE FOR Old:New:Comment:

The field can contain:

[Old Value]:[New Value]:[Comment]

Preserve the complete information.

For example:

[Low Discrimination]:[Shigella boydii]:[]

must be preserved.

If an actual comment exists, preserve the COMPLETE comment.

For example:

Since Mobility test is negative and it is a known standard culture,
Shigella boydii is selected.

must not be removed or shortened.

If all three parts are empty, return an empty string.

Do NOT return:

[][][]
[]:[]:[]
[ ]:[ ]:[ ]

for a completely empty field.

Return ONLY valid JSON.

Use exactly this structure:

{
    "records": [
        {
            "event_date_time": "",
            "accession_number": "",
            "event_type_description": "",
            "event_associated_user_id": "",
            "old_new_comment": "",
            "card_bar_code": ""
        }
    ]
}

Do not use Markdown.
Do not use ```json.
"""

    response = client.models.generate_content(
        model=MODEL,
        contents=[
            {
                "inline_data": {
                    "mime_type": "image/png",
                    "data": image_bytes
                }
            },
            prompt
        ],
        config={
            "response_mime_type": "application/json"
        }
    )

    if not response.text:
        return []

    text = response.text.strip()

    # Remove Markdown fences if Gemini adds them.
    if text.startswith("```"):
        text = text.replace("```json", "")
        text = text.replace("```", "")
        text = text.strip()

    try:
        data = json.loads(text)

    except json.JSONDecodeError:

        print(
            f"WARNING: Could not parse page {page_number}."
        )

        return []

    records = data.get("records", [])

    cleaned_records = []

    for record in records:

        if not isinstance(record, dict):
            continue

        cleaned_record = {

            "event_date_time": clean_text(
                record.get("event_date_time", "")
            ),

            "accession_number": clean_text(
                record.get("accession_number", "")
            ),

            "event_type_description": clean_text(
                record.get("event_type_description", "")
            ),

            "event_associated_user_id": clean_text(
                record.get("event_associated_user_id", "")
            ),

            "old_new_comment": clean_comment(
                record.get("old_new_comment", "")
            ),

            "card_bar_code": clean_text(
                record.get("card_bar_code", "")
            )
        }

        cleaned_records.append(cleaned_record)

    print(
        f"Page {page_number} completed. "
        f"Records found: {len(cleaned_records)}"
    )

    return cleaned_records


# ============================================================
# PROCESS PDF PAGE BY PAGE
# ============================================================

def extract_pdf(pdf_path):

    print("Opening PDF...")

    document = fitz.open(str(pdf_path))

    total_pages = len(document)

    print(f"Total pages: {total_pages}")
    print()

    all_records = []

    for page_index in range(total_pages):

        page_number = page_index + 1

        print("-" * 60)
        print(f"Page {page_number}/{total_pages}")
        print("-" * 60)

        page = document[page_index]

        # Convert PDF page to PNG image.
        pixmap = page.get_pixmap(
            matrix=fitz.Matrix(2, 2),
            alpha=False
        )

        image_bytes = pixmap.tobytes("png")

        try:

            records = extract_page(
                image_bytes,
                page_number
            )

            all_records.extend(records)

        except Exception as error:

            print(
                f"ERROR on page {page_number}: {error}"
            )

        print()

    document.close()

    return all_records


# ============================================================
# CREATE EXCEL FILE
# ============================================================

def create_excel(records, pdf_path):

    output_path = pdf_path.with_suffix(".xlsx")

    print("Creating Excel file...")

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "OCR Data"

    # --------------------------------------------------------
    # Headers
    # --------------------------------------------------------

    headers = [
        "Event Date/Time",
        "Accession Number",
        "Event Type/Description",
        "Event Associated User ID",
        "Old:New:Comment",
        "Card Bar Code"
    ]

    worksheet.append(headers)

    # --------------------------------------------------------
    # Header formatting
    # --------------------------------------------------------

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    for cell in worksheet[1]:

        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # --------------------------------------------------------
    # Add records
    # --------------------------------------------------------

    for record in records:

        worksheet.append([
            record.get("event_date_time", ""),
            record.get("accession_number", ""),
            record.get("event_type_description", ""),
            record.get("event_associated_user_id", ""),
            record.get("old_new_comment", ""),
            record.get("card_bar_code", "")
        ])

    # --------------------------------------------------------
    # Format all data cells
    # --------------------------------------------------------

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    # --------------------------------------------------------
    # Freeze header
    # --------------------------------------------------------

    worksheet.freeze_panes = "A2"

    # --------------------------------------------------------
    # Enable Excel filters
    # --------------------------------------------------------

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # --------------------------------------------------------
    # Set column widths
    # --------------------------------------------------------

    column_widths = {
        "A": 24,
        "B": 25,
        "C": 35,
        "D": 28,
        "E": 60,
        "F": 25
    }

    for column, width in column_widths.items():

        worksheet.column_dimensions[
            column
        ].width = width

    # --------------------------------------------------------
    # Set row height
    # --------------------------------------------------------

    worksheet.row_dimensions[1].height = 35

    # --------------------------------------------------------
    # Save workbook
    # --------------------------------------------------------

    workbook.save(output_path)

    return output_path


# ============================================================
# MAIN
# ============================================================

def main():

    print()
    print("=" * 70)
    print("STARTING OCR")
    print("=" * 70)
    print()

    # --------------------------------------------------------
    # Check command-line argument
    # --------------------------------------------------------

    if len(sys.argv) != 2:

        print("Usage:")
        print()

        print(
            'python ocr.py "C:\\path\\to\\document.pdf"'
        )

        print()

        sys.exit(1)

    # --------------------------------------------------------
    # Get PDF path
    # --------------------------------------------------------

    pdf_path = Path(sys.argv[1])

    # --------------------------------------------------------
    # Check file exists
    # --------------------------------------------------------

    if not pdf_path.exists():

        print("ERROR: PDF file not found.")
        print()
        print("Path:")
        print(pdf_path)

        sys.exit(1)

    # --------------------------------------------------------
    # Check extension
    # --------------------------------------------------------

    if pdf_path.suffix.lower() != ".pdf":

        print("ERROR: Input file must be a PDF.")

        sys.exit(1)

    print(f"PDF: {pdf_path.name}")
    print()

    try:

        # ----------------------------------------------------
        # Extract records
        # ----------------------------------------------------

        records = extract_pdf(pdf_path)

        # ----------------------------------------------------
        # Show total records
        # ----------------------------------------------------

        print()
        print("=" * 70)

        print(
            f"Total records extracted: {len(records)}"
        )

        print("=" * 70)

        # ----------------------------------------------------
        # Create Excel
        # ----------------------------------------------------

        output_path = create_excel(
            records,
            pdf_path
        )

        print()
        print("Excel file created:")
        print(output_path)

        print()
        print("=" * 70)
        print("OCR COMPLETED")
        print("=" * 70)
        print()

    except KeyboardInterrupt:

        print()
        print()
        print("OCR cancelled by user.")

        sys.exit(1)

    except Exception as error:

        print()
        print("=" * 70)
        print("ERROR")
        print("=" * 70)
        print()
        print(error)
        print()

        sys.exit(1)


# ============================================================
# PROGRAM ENTRY POINT
# ============================================================

if __name__ == "__main__":
    main()